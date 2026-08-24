from __future__ import annotations

import os

os.environ["TORCH_COMPILE_DISABLE"] = "1"
os.environ["PYTORCH_JIT"] = "0"
os.environ["OMP_NUM_THREADS"] = "1"

import importlib.util
import logging
from pathlib import Path

import openpyxl
import pandas as pd

from src.extract.base import (
    BaseExtractor,
    DocumentSection,
    ExtractedDocument,
    generate_document_id,
)

logger = logging.getLogger(__name__)
_HAS_DOCLING = importlib.util.find_spec("docling") is not None


class XLSXExtractor(BaseExtractor):
    def extract(self, file_path: Path, role: str, category: str) -> ExtractedDocument:
        if _HAS_DOCLING:
            docling_result = self._try_docling_extract(file_path, role, category)
            if docling_result:
                return docling_result

        return self._pandas_xlsx_extract(file_path, role, category)

    def _try_docling_extract(self, file_path: Path, role: str, category: str) -> ExtractedDocument | None:
        try:
            from docling.document_converter import DocumentConverter

            converter = DocumentConverter()
            res = converter.convert(str(file_path))
            docling_doc = res.document
            markdown_content = docling_doc.export_to_markdown()

            doc_id = generate_document_id(category, file_path.name)
            title = file_path.stem.replace("_", " ").replace("-", " ")

            sections = [
                DocumentSection(
                    title=title,
                    level=1,
                    content=markdown_content,
                    section_type="table",
                )
            ]

            return ExtractedDocument(
                document_id=doc_id,
                title=title,
                source_file=file_path.name,
                source_path=f"{category}/{file_path.name}",
                document_type="xlsx",
                role=role,
                category=category,
                pages=1,
                sections=sections,
                images=[],
                raw_text=markdown_content,
            )
        except BaseException as e:
            logger.debug(f"Docling extraction skipped for XLSX {file_path.name}: {e}")
            return None

    def _pandas_xlsx_extract(self, file_path: Path, role: str, category: str) -> ExtractedDocument:
        sections: list[DocumentSection] = []
        full_raw_text: list[str] = []

        parsed = False
        excel_file = None
        try:
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names

            for sheet_name in sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                df = df.dropna(how="all")

                if df.empty:
                    continue

                try:
                    md_table = df.to_markdown(index=False)
                except Exception:
                    headers = [str(c) for c in df.columns]
                    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
                    for row in df.itertuples(index=False):
                        lines.append("| " + " | ".join(str(c if c is not None else "").replace("\n", " ") for c in row) + " |")
                    md_table = "\n".join(lines)

                if not md_table:
                    continue

                section_content = f"### Worksheet: {sheet_name}\n\n{md_table}"
                full_raw_text.append(section_content)

                sections.append(
                    DocumentSection(
                        title=f"Worksheet: {sheet_name}",
                        level=2,
                        content=section_content,
                        section_type="table",
                        sheet_name=sheet_name,
                    )
                )
            parsed = True
        except Exception as e:
            logger.debug(f"Pandas parse XLSX exception for {file_path}: {e}")
        finally:
            if excel_file is not None:
                excel_file.close()

        if not parsed:
            wb = None
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    rows = [r for r in rows if any(cell is not None for cell in r)]
                    if not rows:
                        continue

                    headers = [str(c or "").strip() for c in rows[0]]
                    md_lines = [
                        "| " + " | ".join(headers) + " |",
                        "| " + " | ".join(["---"] * len(headers)) + " |",
                    ]
                    for r in rows[1:]:
                        vals = [str(c or "").strip().replace("\n", " ") for c in r]
                        md_lines.append("| " + " | ".join(vals) + " |")

                    md_table = "\n".join(md_lines)
                    section_content = f"### Worksheet: {sheet_name}\n\n{md_table}"
                    full_raw_text.append(section_content)

                    sections.append(
                        DocumentSection(
                            title=f"Worksheet: {sheet_name}",
                            level=2,
                            content=section_content,
                            section_type="table",
                            sheet_name=sheet_name,
                        )
                    )
            finally:
                if wb is not None:
                    wb.close()

        doc_id = generate_document_id(category, file_path.name)
        title = file_path.stem.replace("_", " ").replace("-", " ")

        return ExtractedDocument(
            document_id=doc_id,
            title=title,
            source_file=file_path.name,
            source_path=f"{category}/{file_path.name}",
            document_type="xlsx",
            role=role,
            category=category,
            pages=len(sections) or 1,
            sections=sections,
            images=[],
            raw_text="\n\n".join(full_raw_text),
        )

